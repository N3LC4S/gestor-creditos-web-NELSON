import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64

# Configurar la página
st.set_page_config(page_title="Gestión de Créditos", layout="wide")

# Función para calcular próximo pago
def calcular_proximo_pago(fecha_ultimo_pago, tipo_pago):
    if pd.isnull(fecha_ultimo_pago):
        return ""
    dias = {"diario": 1, "semanal": 7, "quincenal": 15, "mensual": 30}
    return fecha_ultimo_pago + timedelta(days=dias.get(tipo_pago, 0))

# Función para calcular estatus
def calcular_estatus(proximo_pago):
    if pd.isnull(proximo_pago):
        return ""
    hoy = datetime.now().date()
    if proximo_pago.date() == hoy:
        return "Pagan hoy"
    elif proximo_pago.date() < hoy:
        return "Vencido"
    else:
        return "Al día"

# Cargar archivo Excel
st.title("📊 Sistema de Gestión de Créditos")
cargado = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

if cargado:
    df_original = pd.read_excel(cargado)
    df_original.columns = df_original.columns.str.strip()  # Normalizar nombres de columnas

    if "Estatus" not in df_original.columns:
        st.error("El archivo debe contener la columna 'Estatus'. Revisa que esté escrita correctamente.")
        st.stop()

    df_original = df_original.fillna("")

    # Calcular automáticamente el estatus y próxima fecha al cargar
    for i, row in df_original.iterrows():
        try:
            tipo_pago = str(row["Frecuencia_pago"]).strip().lower()
            fecha_pago = pd.to_datetime(row["Fecha_ultimo_pago"], errors='coerce')
            proximo = calcular_proximo_pago(fecha_pago, tipo_pago)
            estatus = calcular_estatus(proximo)
            df_original.at[i, "Proxima_fecha_pago"] = proximo
            df_original.at[i, "Estatus"] = estatus
        except:
            continue

    if "data" not in st.session_state:
        st.session_state.data = df_original.copy()

    df = st.session_state.data

    # Botón para agregar nuevo cobro
    if st.button("➕ Agregar nuevo cobro"):
        nueva_fila = pd.Series({
            "Fecha_deuda": datetime.now().date(),
            "Cliente": "",
            "Monto": 0,
            "Frecuencia_pago": "diario",
            "Proxima_fecha_pago": "",
            "Fecha_ultimo_pago": "",
            "Pagos_realizados": 0,
            "Estatus": ""
        })
        st.session_state.data = pd.concat([st.session_state.data, nueva_fila.to_frame().T], ignore_index=True)
        df = st.session_state.data

    # Filtros
    col1, col2 = st.columns(2)
    with col1:
        filtro_nombre = st.text_input("🔍 Buscar por nombre")
    with col2:
        filtro_estado = st.selectbox("📌 Filtrar por estatus", ["Todos"] + sorted(df["Estatus"].unique().tolist()))

    df_filtrado = df.copy()
    if filtro_nombre:
        df_filtrado = df_filtrado[df_filtrado["Cliente"].str.contains(filtro_nombre, case=False, na=False)]
    if filtro_estado != "Todos":
        df_filtrado = df_filtrado[df_filtrado["Estatus"] == filtro_estado]

    # Edición interactiva
    edited_df = st.data_editor(df_filtrado, num_rows="dynamic", use_container_width=True)

    # Procesar edición
    for i, row in edited_df.iterrows():
        idx = df[(df["Cliente"] == row["Cliente"]) & (df["Monto"] == row["Monto"])].index
        if not idx.empty:
            i_real = idx[0]
            df.loc[i_real] = row
            try:
                tipo_pago = row["Frecuencia_pago"].strip().lower()
                fecha_pago = pd.to_datetime(row["Fecha_ultimo_pago"], errors='coerce')
                proximo = calcular_proximo_pago(fecha_pago, tipo_pago)
                estatus = calcular_estatus(proximo)
                df.at[i_real, "Proxima_fecha_pago"] = proximo
                df.at[i_real, "Estatus"] = estatus
            except:
                pass

    st.success("Cambios guardados en memoria.")

    # Botón para descargar
    def descargar_excel(df):
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Créditos')
        ws = writer.book['Créditos']

        for row in range(2, len(df) + 2):
            estado = str(df.loc[row - 2, "Estatus"]).lower()
            color = "FFFFFF"
            if estado == "vencido":
                color = "FF9999"
            elif estado == "pagan hoy":
                color = "FFFF99"
            elif estado == "al día":
                color = "CCFFCC"
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        writer.close()
        output.seek(0)
        return output.getvalue()

    st.download_button("📥 Descargar Excel completo con colores", data=descargar_excel(df), file_name="creditos_actualizados.xlsx")
